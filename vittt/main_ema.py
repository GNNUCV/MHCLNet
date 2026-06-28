import os
import time
import argparse
import datetime
import json
import csv
import numpy as np
import random

import torch
import torch.nn as nn
import torch.backends.cudnn as cudnn
import torch.distributed as dist
from torch.cuda.amp import autocast, GradScaler

from timm.loss import LabelSmoothingCrossEntropy, SoftTargetCrossEntropy
from timm.utils import accuracy, AverageMeter, ModelEma

from config import get_config
from models import build_model
from data import build_loader
from lr_scheduler import build_scheduler
from optimizer import build_optimizer
from logger import create_logger
from utils_ema import (
    load_checkpoint,
    save_checkpoint,
    get_grad_norm,
    auto_resume_helper,
    reduce_tensor,
    load_pretrained,
)

import warnings
warnings.filterwarnings("ignore")

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


def parse_option():
    parser = argparse.ArgumentParser("ViT^3 training and evaluation script", add_help=False)
    parser.add_argument("--cfg", type=str, required=True, metavar="FILE", help="path to config file")
    parser.add_argument(
        "--opts",
        help="Modify config options by adding 'KEY VALUE' pairs.",
        default=None,
        nargs="+",
    )

    parser.add_argument("--batch-size", type=int, help="batch size for single GPU")
    parser.add_argument("--data-path", type=str, help="path to dataset")
    parser.add_argument("--zip", action="store_true", help="use zipped dataset instead of folder dataset")
    parser.add_argument(
        "--cache-mode",
        type=str,
        default="part",
        choices=["no", "full", "part"],
        help="no: no cache, full: cache all data, part: shard dataset and cache one part",
    )
    parser.add_argument("--resume", help="resume from checkpoint")
    parser.add_argument("--use-checkpoint", action="store_true", help="use gradient checkpointing to save memory")
    parser.add_argument("--amp", action="store_true", default=False)
    parser.add_argument(
        "--output",
        default="output",
        type=str,
        metavar="PATH",
        help="root of output folder, full path is <output>/<model_name>/<tag>",
    )
    parser.add_argument("--tag", help="tag of experiment")
    parser.add_argument("--eval", action="store_true", help="Perform evaluation only")
    parser.add_argument("--throughput", action="store_true", help="Test throughput only")
    parser.add_argument("--pretrained", type=str, help="Finetune initial checkpoint.", default="")
    parser.add_argument("--find-unused-params", action="store_true", default=False)

    parser.add_argument("--freeze-backbone", action="store_true", help="Freeze backbone and train selected modules")
    parser.add_argument("--eval-split", type=str, default="val", choices=["val", "test"],
                        help="Evaluation split, choose val or test.")

    # 默认关闭 EMA。你现在不用 EMA，所以不会打印 EMA 验证信息。
    parser.add_argument("--model-ema", dest="model_ema", action="store_true", default=False,
                        help="Enable tracking moving average of model weights")
    parser.add_argument("--no-model-ema", dest="model_ema", action="store_false",
                        help="Disable model EMA")
    parser.add_argument("--model-ema-force-cpu", action="store_true", default=False,
                        help="Force ema to be tracked on CPU, rank=0 node only.")
    parser.add_argument("--model-ema-decay", type=float, default=0.9996,
                        help="decay factor for model weights moving average")

    args, unparsed = parser.parse_known_args()
    config = get_config(args)

    return args, config


def freeze_backbone_train_selected(model, logger, train_keywords=("head", "mhcla")):
    """
    冻结原始骨干，只训练指定关键字包含的模块。
    """
    for name, param in model.named_parameters():
        param.requires_grad = False

    for name, param in model.named_parameters():
        if any(key in name for key in train_keywords):
            param.requires_grad = True

    trainable_params = sum(p.numel() for p in model.parameters() if p.requires_grad)
    total_params = sum(p.numel() for p in model.parameters())

    logger.info("Freeze backbone and train selected modules.")
    logger.info(f"Train keywords: {train_keywords}")
    logger.info(f"Trainable params: {trainable_params / 1e6:.4f}M / {total_params / 1e6:.4f}M")

    for name, param in model.named_parameters():
        if param.requires_grad:
            logger.info(f"Trainable parameter: {name}")


def _rankdata_average(values):
    """
    Tie-aware average ranks, 1-based.
    用于无 sklearn 依赖地计算 AUC。
    """
    values = np.asarray(values)
    order = np.argsort(values)
    ranks = np.empty(len(values), dtype=np.float64)

    i = 0
    while i < len(values):
        j = i
        while j + 1 < len(values) and values[order[j + 1]] == values[order[i]]:
            j += 1
        avg_rank = (i + 1 + j + 1) / 2.0
        ranks[order[i:j + 1]] = avg_rank
        i = j + 1

    return ranks


def _binary_auc(y_true_binary, scores):
    y_true_binary = np.asarray(y_true_binary).astype(np.int64)
    scores = np.asarray(scores, dtype=np.float64)

    n_pos = int((y_true_binary == 1).sum())
    n_neg = int((y_true_binary == 0).sum())

    if n_pos == 0 or n_neg == 0:
        return np.nan

    ranks = _rankdata_average(scores)
    pos_rank_sum = ranks[y_true_binary == 1].sum()
    auc = (pos_rank_sum - n_pos * (n_pos + 1) / 2.0) / (n_pos * n_neg)

    return float(auc)


def build_confusion_matrix(y_true, y_pred, num_classes):
    cm = np.zeros((num_classes, num_classes), dtype=np.int64)
    for t, p in zip(y_true, y_pred):
        cm[int(t), int(p)] += 1
    return cm


def safe_div(a, b):
    return a / b if b != 0 else 0.0


def compute_classification_metrics(y_true, y_pred, y_prob, num_classes):
    """
    计算 Accuracy、Macro Precision、Macro Recall、Macro F1、Macro AUC 和混淆矩阵。
    AUC 使用 one-vs-rest macro AUC。
    如果某一类在当前 split 中没有正样本或负样本，则该类 AUC 记为 NaN，不参与 macro 平均。
    """
    y_true = np.asarray(y_true, dtype=np.int64)
    y_pred = np.asarray(y_pred, dtype=np.int64)
    y_prob = np.asarray(y_prob, dtype=np.float64)

    cm = build_confusion_matrix(y_true, y_pred, num_classes)

    total = cm.sum()
    correct = np.trace(cm)
    acc = safe_div(correct, total) * 100.0

    per_class = []
    precisions = []
    recalls = []
    f1s = []
    aucs = []

    for i in range(num_classes):
        tp = cm[i, i]
        fp = cm[:, i].sum() - tp
        fn = cm[i, :].sum() - tp
        support = cm[i, :].sum()

        precision = safe_div(tp, tp + fp)
        recall = safe_div(tp, tp + fn)
        f1 = safe_div(2 * precision * recall, precision + recall)

        if y_prob.ndim == 2 and y_prob.shape[1] > i:
            auc_i = _binary_auc((y_true == i).astype(np.int64), y_prob[:, i])
        else:
            auc_i = np.nan

        precisions.append(precision)
        recalls.append(recall)
        f1s.append(f1)
        aucs.append(auc_i)

        per_class.append({
            "class_index": i,
            "precision": precision * 100.0,
            "recall": recall * 100.0,
            "f1": f1 * 100.0,
            "auc": None if np.isnan(auc_i) else float(auc_i),
            "support": int(support),
        })

    valid_aucs = [a for a in aucs if not np.isnan(a)]
    macro_auc = float(np.mean(valid_aucs)) if len(valid_aucs) > 0 else np.nan

    return {
        "accuracy": float(acc),
        "macro_precision": float(np.mean(precisions) * 100.0),
        "macro_recall": float(np.mean(recalls) * 100.0),
        "macro_f1": float(np.mean(f1s) * 100.0),
        "macro_auc": macro_auc,
        "per_class": per_class,
        "confusion_matrix": cm,
    }


def gather_numpy_arrays(local_array):
    """
    单卡时直接返回。
    多卡时用 all_gather_object 收集不同长度的 numpy 数组。
    """
    local_array = np.asarray(local_array)

    if not (dist.is_available() and dist.is_initialized()) or dist.get_world_size() == 1:
        return local_array

    gathered = [None for _ in range(dist.get_world_size())]
    dist.all_gather_object(gathered, local_array)

    if local_array.ndim == 0:
        return np.asarray(gathered)

    return np.concatenate(gathered, axis=0)


def matrix_to_string(cm):
    cm = np.asarray(cm, dtype=np.int64)
    return np.array2string(cm, separator=", ")


def append_metrics_to_excel(excel_path, record, logger=None):
    """
    把每个 epoch 的指标追加到 Excel。
    如果没有 openpyxl，则自动写入同名 csv 作为兜底。
    """
    columns = [
        "epoch",
        "phase",
        "train_loss",
        "val_loss",
        "acc1",
        "acc5",
        "accuracy",
        "macro_precision",
        "macro_recall",
        "macro_f1",
        "macro_auc",
        "confusion_matrix",
        "per_class_metrics",
        "lr",
    ]

    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    if Workbook is None or load_workbook is None:
        csv_path = os.path.splitext(excel_path)[0] + ".csv"
        file_exists = os.path.exists(csv_path)
        with open(csv_path, "a", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            if not file_exists:
                writer.writeheader()
            writer.writerow({col: record.get(col, "") for col in columns})
        if logger is not None:
            logger.warning(f"openpyxl is not installed. Metrics were saved to CSV: {csv_path}")
        return

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "metrics"
        ws.append(columns)

    ws.append([record.get(col, "") for col in columns])
    wb.save(excel_path)


def log_metrics_summary(logger, epoch, phase, metrics, train_loss=None):
    auc_value = metrics.get("macro_auc", np.nan)
    auc_text = "nan" if np.isnan(auc_value) else f"{auc_value:.4f}"

    if str(phase).startswith("test"):
        logger.info(
            f"{phase} summary | "
            f"Acc@1 {metrics['acc1']:.2f}% | "
            f"Acc@5 {metrics['acc5']:.2f}% | "
            f"Accuracy {metrics['accuracy']:.2f}% | "
            f"Precision {metrics['macro_precision']:.2f}% | "
            f"Recall {metrics['macro_recall']:.2f}% | "
            f"F1 {metrics['macro_f1']:.2f}% | "
            f"AUC {auc_text}"
        )
    elif train_loss is not None:
        logger.info(
            f"Epoch {epoch} {phase} summary | "
            f"Train Loss {train_loss:.4f} | "
            f"Val Loss {metrics['loss']:.4f} | "
            f"Acc@1 {metrics['acc1']:.2f}% | "
            f"Acc@5 {metrics['acc5']:.2f}% | "
            f"Accuracy {metrics['accuracy']:.2f}% | "
            f"Precision {metrics['macro_precision']:.2f}% | "
            f"Recall {metrics['macro_recall']:.2f}% | "
            f"F1 {metrics['macro_f1']:.2f}% | "
            f"AUC {auc_text}"
        )
    else:
        logger.info(
            f"{phase} summary | "
            f"Val Loss {metrics['loss']:.4f} | "
            f"Acc@1 {metrics['acc1']:.2f}% | "
            f"Acc@5 {metrics['acc5']:.2f}% | "
            f"Accuracy {metrics['accuracy']:.2f}% | "
            f"Precision {metrics['macro_precision']:.2f}% | "
            f"Recall {metrics['macro_recall']:.2f}% | "
            f"F1 {metrics['macro_f1']:.2f}% | "
            f"AUC {auc_text}"
        )

    logger.info(
        f"{phase} Confusion matrix rows=true cols=pred:\n"
        f"{matrix_to_string(metrics['confusion_matrix'])}"
    )


def make_excel_record(epoch, phase, metrics, train_loss=None, lr=None):
    record = {
        "epoch": epoch,
        "phase": phase,
        "train_loss": "" if train_loss is None else float(train_loss),
        "val_loss": float(metrics["loss"]) if phase == "val" else "",
        "acc1": float(metrics["acc1"]),
        "acc5": float(metrics["acc5"]),
        "accuracy": float(metrics["accuracy"]),
        "macro_precision": float(metrics["macro_precision"]),
        "macro_recall": float(metrics["macro_recall"]),
        "macro_f1": float(metrics["macro_f1"]),
        "macro_auc": "" if np.isnan(metrics["macro_auc"]) else float(metrics["macro_auc"]),
        "confusion_matrix": matrix_to_string(metrics["confusion_matrix"]),
        "per_class_metrics": json.dumps(metrics["per_class"], ensure_ascii=False),
        "lr": "" if lr is None else float(lr),
    }
    return record


def main():
    os.environ["NCCL_BLOCKING_WAIT"] = "1"
    args, config = parse_option()

    rank = int(os.environ["RANK"])
    world_size = int(os.environ["WORLD_SIZE"])
    local_rank = int(os.environ["LOCAL_RANK"])
    torch.cuda.set_device(local_rank)
    dist.init_process_group(backend="nccl", init_method="env://", world_size=world_size, rank=rank)

    # seed = config.SEED + dist.get_rank()
    # random.seed(seed)
    # np.random.seed(seed)
    # torch.manual_seed(seed)
    # torch.cuda.manual_seed(seed)
    # torch.cuda.manual_seed_all(seed)
    #
    # cudnn.enabled = True
    # cudnn.benchmark = True
    # cudnn.deterministic = True

    seed = config.SEED + dist.get_rank()
    torch.manual_seed(seed)
    np.random.seed(seed)
    cudnn.enabled = True
    cudnn.benchmark = True

    linear_scaled_lr = config.TRAIN.BASE_LR * config.DATA.BATCH_SIZE * dist.get_world_size() / 512.0
    linear_scaled_warmup_lr = config.TRAIN.WARMUP_LR * config.DATA.BATCH_SIZE * dist.get_world_size() / 512.0
    linear_scaled_min_lr = config.TRAIN.MIN_LR * config.DATA.BATCH_SIZE * dist.get_world_size() / 512.0

    config.defrost()
    config.TRAIN.BASE_LR = linear_scaled_lr
    config.TRAIN.WARMUP_LR = linear_scaled_warmup_lr
    config.TRAIN.MIN_LR = linear_scaled_min_lr
    config.LOCAL_RANK = local_rank
    config.freeze()

    if args.model_ema:
        args.model_ema_decay = args.model_ema_decay ** (config.DATA.BATCH_SIZE * dist.get_world_size() / 4096.0)

    os.makedirs(config.OUTPUT, exist_ok=True)
    logger = create_logger(output_dir=config.OUTPUT, dist_rank=dist.get_rank(), name=f"{config.MODEL.NAME}")

    if dist.get_rank() == 0:
        path = os.path.join(config.OUTPUT, "config.json")
        with open(path, "w") as f:
            f.write(config.dump())
        logger.info(f"Full config saved to {path}")

    logger.info(config.dump())

    _, dataset_val, dataset_test, data_loader_train, data_loader_val, data_loader_test, mixup_fn = build_loader(config)

    logger.info(f"Creating model:{config.MODEL.TYPE}/{config.MODEL.NAME}")
    model = build_model(config)

    if args.freeze_backbone:
        freeze_backbone_train_selected(
            model,
            logger,
            train_keywords=("head", "mhcla")
        )

    model.cuda()
    logger.info(str(model))

    optimizer = build_optimizer(config, model)

    model = nn.parallel.DistributedDataParallel(
        model,
        device_ids=[local_rank],
        broadcast_buffers=True,
        find_unused_parameters=args.find_unused_params,
    )
    model_without_ddp = model.module

    lr_scheduler = build_scheduler(config, optimizer, len(data_loader_train))
    total_epochs = config.TRAIN.EPOCHS + config.TRAIN.COOLDOWN_EPOCHS

    if config.AUG.MIXUP > 0.:
        criterion = SoftTargetCrossEntropy()
    elif config.MODEL.LABEL_SMOOTHING > 0.:
        criterion = LabelSmoothingCrossEntropy(smoothing=config.MODEL.LABEL_SMOOTHING)
    else:
        criterion = nn.CrossEntropyLoss()

    max_accuracy = 0.0
    max_accuracy_e = 0.0

    metrics_excel_path = os.path.join(config.OUTPUT, "epoch_metrics.xlsx")

    if args.pretrained != "":
        load_pretrained(args.pretrained, model_without_ddp, logger)

    if config.TRAIN.AUTO_RESUME:
        resume_file = auto_resume_helper(config.OUTPUT)
        if resume_file:
            if config.MODEL.RESUME:
                logger.warning(f"auto-resume changing resume file from {config.MODEL.RESUME} to {resume_file}")
            config.defrost()
            config.MODEL.RESUME = resume_file
            config.freeze()
            logger.info(f"auto resuming from {resume_file}")
        else:
            logger.info(f"no checkpoint found in {config.OUTPUT}, ignoring auto resume")

    if config.MODEL.RESUME:
        max_accuracy, max_accuracy_e = load_checkpoint(
            config, model_without_ddp, optimizer, lr_scheduler, logger
        )

    if config.EVAL_MODE:
        if args.eval_split == "test":
            if data_loader_test is None:
                raise RuntimeError(
                    f"--eval-split test was used, but test folder was not found under {config.DATA.DATA_PATH}"
                )
            eval_loader = data_loader_test
            split_name = "test"
        else:
            eval_loader = data_loader_val
            split_name = "val"

        eval_metrics = validate(config, eval_loader, model, logger, split_name=split_name)
        log_metrics_summary(logger, "eval", split_name, eval_metrics, train_loss=None)

        if dist.get_rank() == 0:
            append_metrics_to_excel(
                metrics_excel_path,
                make_excel_record("eval", split_name, eval_metrics, train_loss=None, lr=None),
                logger=logger,
            )

        return

    model_ema = None
    if args.model_ema:
        model_ema = ModelEma(
            model,
            decay=args.model_ema_decay,
            device="cpu" if args.model_ema_force_cpu else "",
            resume=config.MODEL.RESUME,
        )

    if config.THROUGHPUT_MODE:
        throughput(data_loader_val, model, logger)
        return

    logger.info("Start training")
    start_time = time.time()

    for epoch in range(config.TRAIN.START_EPOCH, total_epochs):
        data_loader_train.sampler.set_epoch(epoch)

        mesa_weight = config.TRAIN.MESA if (model_ema is not None and epoch >= int(0.25 * total_epochs)) else -1.0

        train_loss = train_one_epoch(
            config,
            model,
            model_ema,
            criterion,
            data_loader_train,
            optimizer,
            epoch,
            mixup_fn,
            lr_scheduler,
            logger,
            total_epochs,
            mesa=mesa_weight,
        )

        val_metrics = validate(config, data_loader_val, model, logger, split_name="val")
        current_lr = optimizer.param_groups[0]["lr"]

        log_metrics_summary(logger, epoch + 1, "val", val_metrics, train_loss=train_loss)

        if dist.get_rank() == 0:
            append_metrics_to_excel(
                metrics_excel_path,
                make_excel_record(epoch + 1, "val", val_metrics, train_loss=train_loss, lr=current_lr),
                logger=logger,
            )

        acc1 = val_metrics["acc1"]
        acc1_e = 0.0

        if dist.get_rank() == 0 and ((epoch + 1) % config.SAVE_FREQ == 0 or (epoch + 1) == total_epochs):
            save_checkpoint(
                config,
                epoch + 1,
                model_without_ddp,
                model_ema,
                max(max_accuracy, acc1),
                max_accuracy_e,
                optimizer,
                lr_scheduler,
                logger,
            )

        if dist.get_rank() == 0 and acc1 >= max_accuracy:
            save_checkpoint(
                config,
                epoch + 1,
                model_without_ddp,
                model_ema,
                max(max_accuracy, acc1),
                max_accuracy_e,
                optimizer,
                lr_scheduler,
                logger,
                name="max_acc",
            )

        max_accuracy = max(max_accuracy, acc1)
        logger.info(f"Max validation accuracy: {max_accuracy:.2f}%")

        if model_ema is not None:
            max_accuracy_e = max(max_accuracy_e, acc1_e)

    dist.barrier()

    if data_loader_test is None:
        logger.warning(
            f"Test folder was not found under {config.DATA.DATA_PATH}. Skip final test."
        )
    else:
        best_ckpt_path = os.path.join(config.OUTPUT, "max_acc.pth")

        if os.path.exists(best_ckpt_path):
            checkpoint = torch.load(best_ckpt_path, map_location="cpu")
            msg = model_without_ddp.load_state_dict(checkpoint["model"], strict=False)
            logger.info(f"Loaded best validation checkpoint from {best_ckpt_path}")
            logger.info(msg)

            test_metrics = validate(config, data_loader_test, model, logger, split_name="test")
            log_metrics_summary(logger, "best_val", "test_best_val", test_metrics, train_loss=None)

            if dist.get_rank() == 0:
                append_metrics_to_excel(
                    metrics_excel_path,
                    make_excel_record("best_val", "test_best_val", test_metrics, train_loss=None, lr=None),
                    logger=logger,
                )
        else:
            logger.warning(f"Best validation checkpoint not found: {best_ckpt_path}")

    dist.barrier()

    total_time = time.time() - start_time
    total_time_str = str(datetime.timedelta(seconds=int(total_time)))
    logger.info("Training time {}".format(total_time_str))


def train_one_epoch(config, model, model_ema, criterion, data_loader, optimizer, epoch, mixup_fn,
                    lr_scheduler, logger, total_epochs, mesa=1.0):
    model.train()
    optimizer.zero_grad()

    num_steps = len(data_loader)
    batch_time = AverageMeter()
    loss_meter = AverageMeter()
    norm_meter = AverageMeter()

    start = time.time()
    end = time.time()

    scaler = GradScaler()

    for idx, (samples, targets) in enumerate(data_loader):
        optimizer.zero_grad()
        samples = samples.cuda(non_blocking=True)
        targets = targets.cuda(non_blocking=True)

        if mixup_fn is not None:
            samples, targets = mixup_fn(samples, targets)

        if config.AMP:
            with autocast():
                if mesa > 0 and model_ema is not None:
                    with torch.inference_mode():
                        ema_output = model_ema.ema(samples).detach()
                    ema_output = torch.clone(ema_output)
                    ema_output = ema_output.softmax(dim=-1).detach()
                    outputs = model(samples)
                    loss = criterion(outputs, targets) + criterion(outputs, ema_output) * mesa
                else:
                    outputs = model(samples)
                    loss = criterion(outputs, targets)

            scaler.scale(loss).backward()
            if config.TRAIN.CLIP_GRAD:
                scaler.unscale_(optimizer)
                grad_norm = nn.utils.clip_grad_norm_(model.parameters(), config.TRAIN.CLIP_GRAD)
                scaler.step(optimizer)
                scaler.update()
            else:
                grad_norm = get_grad_norm(model.parameters())
                scaler.step(optimizer)
                scaler.update()
        else:
            if mesa > 0 and model_ema is not None:
                with torch.inference_mode():
                    ema_output = model_ema.ema(samples).detach()
                ema_output = torch.clone(ema_output)
                ema_output = ema_output.softmax(dim=-1).detach()
                outputs = model(samples)
                loss = criterion(outputs, targets) + criterion(outputs, ema_output) * mesa
            else:
                outputs = model(samples)
                loss = criterion(outputs, targets)

            loss.backward()
            if config.TRAIN.CLIP_GRAD:
                grad_norm = nn.utils.clip_grad_norm_(model.parameters(), config.TRAIN.CLIP_GRAD)
            else:
                grad_norm = get_grad_norm(model.parameters())
            optimizer.step()

        lr_scheduler.step_update(epoch * num_steps + idx)

        torch.cuda.synchronize()

        if model_ema is not None:
            model_ema.update(model)

        loss_meter.update(loss.item(), samples.size(0))
        norm_meter.update(grad_norm)
        batch_time.update(time.time() - end)
        end = time.time()

        if (idx + 1) % config.PRINT_FREQ == 0:
            lr = optimizer.param_groups[0]["lr"]
            memory_used = torch.cuda.max_memory_allocated() / (1024.0 * 1024.0)
            etas = batch_time.avg * (num_steps - idx)
            logger.info(
                f"Train: [{epoch + 1}/{total_epochs}][{idx + 1}/{num_steps}]\t"
                f"eta {datetime.timedelta(seconds=int(etas))} lr {lr:.6f}\t"
                f"time {batch_time.val:.4f} ({batch_time.avg:.4f})\t"
                f"loss {loss_meter.val:.4f} ({loss_meter.avg:.4f})\t"
                f"grad_norm {norm_meter.val:.4f} ({norm_meter.avg:.4f})\t"
                f"mem {memory_used:.0f}MB"
            )

    epoch_time = time.time() - start
    logger.info(f"EPOCH {epoch + 1} training takes {datetime.timedelta(seconds=int(epoch_time))}")
    logger.info(f"EPOCH {epoch + 1} train loss: {loss_meter.avg:.4f}")

    return loss_meter.avg


@torch.no_grad()
def validate(config, data_loader, model, logger, split_name="val"):
    criterion = nn.CrossEntropyLoss()
    model.eval()

    batch_time = AverageMeter()
    loss_meter = AverageMeter()
    acc1_meter = AverageMeter()
    acc5_meter = AverageMeter()

    y_true_list = []
    y_pred_list = []
    y_prob_list = []

    end = time.time()

    for idx, (images, target) in enumerate(data_loader):
        images = images.cuda(non_blocking=True)
        target = target.cuda(non_blocking=True)

        output = model(images)

        loss = criterion(output, target)

        if output.shape[1] >= 5:
            acc1, acc5 = accuracy(output, target, topk=(1, 5))
        else:
            acc1 = accuracy(output, target, topk=(1,))[0]
            acc5 = acc1

        prob = torch.softmax(output, dim=1)
        pred = torch.argmax(prob, dim=1)

        y_true_list.append(target.detach().cpu().numpy())
        y_pred_list.append(pred.detach().cpu().numpy())
        y_prob_list.append(prob.detach().cpu().numpy())

        acc1 = reduce_tensor(acc1)
        acc5 = reduce_tensor(acc5)
        loss = reduce_tensor(loss)

        loss_meter.update(loss.item(), target.size(0))
        acc1_meter.update(acc1.item(), target.size(0))
        acc5_meter.update(acc5.item(), target.size(0))

        batch_time.update(time.time() - end)
        end = time.time()

        if (idx + 1) % config.PRINT_FREQ == 0:
            memory_used = torch.cuda.max_memory_allocated() / (1024.0 * 1024.0)
            logger.info(
                f"{split_name}: [{idx + 1}/{len(data_loader)}]\t"
                f"Time {batch_time.val:.3f} ({batch_time.avg:.3f})\t"
                f"Loss {loss_meter.val:.4f} ({loss_meter.avg:.4f})\t"
                f"Acc@1 {acc1_meter.val:.3f} ({acc1_meter.avg:.3f})\t"
                f"Acc@5 {acc5_meter.val:.3f} ({acc5_meter.avg:.3f})\t"
                f"Mem {memory_used:.0f}MB"
            )

    local_y_true = np.concatenate(y_true_list, axis=0) if len(y_true_list) > 0 else np.array([], dtype=np.int64)
    local_y_pred = np.concatenate(y_pred_list, axis=0) if len(y_pred_list) > 0 else np.array([], dtype=np.int64)
    local_y_prob = np.concatenate(y_prob_list, axis=0) if len(y_prob_list) > 0 else np.empty((0, 0), dtype=np.float64)

    y_true = gather_numpy_arrays(local_y_true)
    y_pred = gather_numpy_arrays(local_y_pred)
    y_prob = gather_numpy_arrays(local_y_prob)

    num_classes = y_prob.shape[1] if y_prob.ndim == 2 and y_prob.shape[1] > 0 else int(config.MODEL.NUM_CLASSES)
    extra_metrics = compute_classification_metrics(y_true, y_pred, y_prob, num_classes)

    metrics = {
        "loss": float(loss_meter.avg),
        "acc1": float(acc1_meter.avg),
        "acc5": float(acc5_meter.avg),
        "accuracy": extra_metrics["accuracy"],
        "macro_precision": extra_metrics["macro_precision"],
        "macro_recall": extra_metrics["macro_recall"],
        "macro_f1": extra_metrics["macro_f1"],
        "macro_auc": extra_metrics["macro_auc"],
        "confusion_matrix": extra_metrics["confusion_matrix"],
        "per_class": extra_metrics["per_class"],
    }

    auc_value = metrics["macro_auc"]
    auc_text = "nan" if np.isnan(auc_value) else f"{auc_value:.4f}"

    if str(split_name).startswith("test"):
        logger.info(
            f"* {split_name} "
            f"Acc@1 {metrics['acc1']:.3f} Acc@5 {metrics['acc5']:.3f} "
            f"Accuracy {metrics['accuracy']:.3f} "
            f"Precision {metrics['macro_precision']:.3f} "
            f"Recall {metrics['macro_recall']:.3f} "
            f"F1 {metrics['macro_f1']:.3f} "
            f"AUC {auc_text}"
        )
    else:
        logger.info(
            f"* {split_name} Loss {metrics['loss']:.4f} "
            f"Acc@1 {metrics['acc1']:.3f} Acc@5 {metrics['acc5']:.3f} "
            f"Accuracy {metrics['accuracy']:.3f} "
            f"Precision {metrics['macro_precision']:.3f} "
            f"Recall {metrics['macro_recall']:.3f} "
            f"F1 {metrics['macro_f1']:.3f} "
            f"AUC {auc_text}"
        )

    logger.info(
        f"* {split_name} Confusion matrix rows=true cols=pred:\n"
        f"{matrix_to_string(metrics['confusion_matrix'])}"
    )

    return metrics


@torch.no_grad()
def throughput(data_loader, model, logger):
    model.eval()

    for _, (images, _) in enumerate(data_loader):
        images = images.cuda(non_blocking=True)
        batch_size = images.shape[0]
        for i in range(50):
            model(images)
        torch.cuda.synchronize()
        logger.info("throughput averaged with 30 times")
        tic1 = time.time()
        for i in range(30):
            model(images)
        torch.cuda.synchronize()
        tic2 = time.time()
        logger.info(f"batch_size {batch_size} throughput {30 * batch_size / (tic2 - tic1)}")
        return


if __name__ == "__main__":
    main()
